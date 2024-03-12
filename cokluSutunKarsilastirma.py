import openpyxl as op
from openpyxl.styles import Font, PatternFill

# Excel dosyalarının yolları
dosya1 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi_2.xlsx')   
sheet1 = dosya1.active
dosya2 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi_1.xlsx')
sheet2 = dosya2.active
dosya3 = op.load_workbook('5GHz_Kapali_Cihaz_Listesi_3.xlsx')
sheet3 = dosya3.active
dosya4 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi_4.xlsx')
sheet4 = dosya4.active

# Yeni Excel dosyası oluştur
output = op.Workbook()
outputSheet = output.active

# A, B, C ve D sütunlarına 'MAC1', 'MAC2', 'MAC3', 'MAC4', 'EŞLEŞME', 'MAC1 Index', 'MAC2 Index', 'MAC3 Index', 'MAC4 Index' yazma
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
column_titles = ['MAC1', 'MAC2', 'MAC3', 'MAC4', 'EŞLEŞME', 'MAC1 Index', 'MAC2 Index', 'MAC3 Index', 'MAC4 Index']
for col, title in zip(columns, column_titles):
    cell = outputSheet[col + '1']
    cell.value = title
    cell.font = Font(bold=True, size=15)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    outputSheet.column_dimensions[col].width = max(sheet1.column_dimensions[col].width, len(title)) + 15

# MAC1, MAC2, MAC3 ve MAC4 verilerini A, B, C ve D sütunlarına yazma
mac1_values = [value.value for value in sheet1['G'][1:]]
mac2_values = [value.value for value in sheet2['G'][1:]]
mac3_values = [value.value for value in sheet3['G'][1:]]
mac4_values = [value.value for value in sheet4['G'][1:]]

# A, B, C ve D sütunlarına MAC1, MAC2, MAC3 ve MAC4 verilerini yazma
for index, (value1, value2, value3, value4) in enumerate(zip(mac1_values, mac2_values, mac3_values, mac4_values), start=2):
    outputSheet['A' + str(index)].value = value1
    outputSheet['B' + str(index)].value = value2
    outputSheet['C' + str(index)].value = value3
    outputSheet['D' + str(index)].value = value4

# Eşleşen başlıkları ve değerleri bulma
common_mac_values = set(filter(None, mac1_values)).intersection(filter(None, mac2_values), filter(None, mac3_values), filter(None, mac4_values))
for index, value in enumerate(common_mac_values, start=1):
    outputSheet.cell(row=index + 1, column=5, value=value)
    mac1_row = mac1_values.index(value) + 2
    mac2_row = mac2_values.index(value) + 2
    mac3_row = mac3_values.index(value) + 2
    mac4_row = mac4_values.index(value) + 2
    outputSheet.cell(row=index + 1, column=6, value=mac1_row)
    outputSheet.cell(row=index + 1, column=7, value=mac2_row)
    outputSheet.cell(row=index + 1, column=8, value=mac3_row)
    outputSheet.cell(row=index + 1, column=9, value=mac4_row)

# Dosyayı kaydet
output.save('tryExcel1.xlsx')
