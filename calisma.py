import openpyxl as op

# Excel dosyasının yolu
dosya1 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi.xlsx')   
sheet1=dosya1.active
dosya2 = op.load_workbook('30_Gun_Uzerinde_Resetlenmemis_Cihaz_Listesi.xlsx')
sheet2=dosya2.active

output=op.load_workbook('tryExcel.xlsx') 
outputSheet=output.active                               
  
# Veri çerçevesinde belirli bir sütunu seçtik
macColumn1=sheet1['G']
macColumn1_List=[]
uzunlukVeri1=0
uzunluk1=len(macColumn1)
for row in macColumn1:
    uzunlukVeri1=uzunlukVeri1+1
    print("veri1 "+ str(uzunlukVeri1) + "/" + str(uzunluk1))
    macColumn1_List.append(row.value)

# Veri çerçevesinde belirli bir sütunu seçtik
macColumn2=sheet2['I']
macColumn2_List=[]
uzunlukVeri2=0
uzunluk2=len(macColumn2)
for row in macColumn2:
    uzunlukVeri2=uzunlukVeri2+1
    print("veri2 "+ str(uzunlukVeri2) + "/" + str(uzunluk2))
    macColumn2_List.append(row.value)


compare_list=[]
uzunluk=len(macColumn1_List)
basla=0
for i in macColumn1_List:
    basla=basla+1
    print(str(basla)+ "/" +str(uzunluk))
    if i in macColumn2_List:
        if i not in compare_list:
            compare_list.append(i)

print(compare_list)
excel_l=range(len(compare_list))
        

for i, j in zip(compare_list, excel_l):
    c1 = outputSheet.cell(row = j+1, column = 1) 
    c1.value = i

output.save('tryExcel.xlsx')