import openpyxl as op
from openpyxl.styles import Font, PatternFill

# Excel dosyalarının yolları
dosya1 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi_2.xlsx')   
sheet1 = dosya1.active
dosya2 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi_1.xlsx')
sheet2 = dosya2.active


# Yeni Excel dosyası oluştur
output = op.Workbook()
outputSheet = output.active

# A, B, C ve D sütunlarına 'MAC1', 'MAC2', 'EŞLEŞME', 'MAC1 Index' ve 'MAC2 Index' yazma
columns = ['A', 'B', 'C', 'D', 'E']
column_titles = ['MAC1', 'MAC2', 'EŞLEŞME', 'MAC1 Index', 'MAC2 Index']
for col, title in zip(columns, column_titles):
    cell = outputSheet[col + '1']
    cell.value = title
    cell.font = Font(bold=True, size=15)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    outputSheet.column_dimensions[col].width = max(sheet1.column_dimensions[col].width, len(title)) + 15

# MAC1 ve MAC2 verilerini A ve B sütunlarına yazma
mac1_values = [value.value for value in sheet1['G'][1:]]
mac2_values = [value.value for value in sheet2['G'][1:]]

# A sütununa MAC1 verilerini yazma
for index, value in enumerate(mac1_values, start=2):
    outputSheet['A' + str(index)].value = value

# B sütununa MAC2 verilerini yazma
for index, value in enumerate(mac2_values, start=2):
    outputSheet['B' + str(index)].value = value

# Eşleşen başlıkları ve değerleri bulma
common_mac_values = set(filter(None, mac1_values)).intersection(filter(None, mac2_values))
for index, value in enumerate(common_mac_values, start=1):
    outputSheet.cell(row=index + 1, column=3, value=value)
    mac1_row = mac1_values.index(value) + 2
    mac2_row = mac2_values.index(value) + 2
    outputSheet.cell(row=index + 1, column=4, value=mac1_row)
    outputSheet.cell(row=index + 1, column=5, value=mac2_row)

# Dosyayı kaydet
output.save('tryExcel1.xlsx')
