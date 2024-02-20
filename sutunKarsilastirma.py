import openpyxl as op
from openpyxl.styles import Font, PatternFill

# Excel dosyalarının yolları
dosya1 = op.load_workbook('5Ghz_Kapali_Cihaz_Listesi.xlsx')   
sheet1 = dosya1.active
dosya2 = op.load_workbook('30_Gun_Uzerinde_Resetlenmemis_Cihaz_Listesi.xlsx')
sheet2 = dosya2.active

# Yeni Excel dosyası oluştur
output = op.Workbook()
outputSheet = output.active

# A, B, C ve D sütunlarına 'MAC1', 'MAC2', 'EŞLEŞME', 'MAC1 Index' ve 'MAC2 Index' yazma
columns = ['A', 'B', 'C', 'D', 'E']
column_titles = ['MAC1', 'MAC2', 'EŞLEŞME', 'MAC1 Index', 'MAC2 Index']
for col, title in zip(columns, column_titles):
    cell = outputSheet[col + '1']
    cell.value = title
    cell.font = Font(bold=True, size=15)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    outputSheet.column_dimensions[col].width = max(sheet1.column_dimensions[col].width, len(title)) + 15

# MAC1 ve MAC2 verilerini A ve B sütunlarına yazma
for index, value in enumerate(sheet1['G'], start=2):
    outputSheet.cell(row=index-1, column=1, value=value.value)

for index, value in enumerate(sheet2['I'], start=2):
    outputSheet.cell(row=index-1, column=2, value=value.value)

# Eşleşme başlığını ve eşleşen değerleri bulma
compare_set = set([cell.value for cell in sheet1['G']])
compare_set.intersection_update([cell.value for cell in sheet2['I']])
for index, value in enumerate(compare_set, start=1):
    outputSheet.cell(row=index + 1, column=3, value=value)
    mac1_row = [i for i, cell in enumerate(sheet1['G'], start=1) if cell.value == value]
    mac2_row = [i for i, cell in enumerate(sheet2['I'], start=1) if cell.value == value]
    outputSheet.cell(row=index + 1, column=4, value=mac1_row[0] if mac1_row else None)
    outputSheet.cell(row=index + 1, column=5, value=mac2_row[0] if mac2_row else None)

# Başlık satırlarının düzenlenmesi
for col, title in zip(columns, column_titles):
    cell = outputSheet[col + '1']
    cell.value = title
    cell.font = Font(bold=True, size=15)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    outputSheet.column_dimensions[col].width = max(sheet1.column_dimensions[col].width, len(title)) + 15

# Dosyayı kaydet
output.save('tryExcel.xlsx')
